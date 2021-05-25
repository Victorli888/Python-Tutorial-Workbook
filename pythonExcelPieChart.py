# import os
# cwd = os.getcwd()
# os.chdir("/path/to/your/folder")
# os.listdir('.')


"""
Consider the following commands to organize, and set up the workspace to run your python code in.
Will help other people using your session to know what directory you're working in
"""


"""Using openpyxl
Lets create a Pie-chart of our spending this month using openpyxl

Notes for this module Located Here: 
https://docs.google.com/document/d/17qukXg7sXoUVJMo9r6UsX_AMUopPopr75qnEFzVLdXo/edit?usp=sharing
"""

import openpyxl as xl
from openpyxl.chart import(
    PieChart,
    Reference)

wb = xl.load_workbook('pythonExcel.xlsx')
sheet1 = wb["Sheet1"]

subTotalHeader = sheet1.cell(1, 4)
subTotalHeader.value = "Sub-Total"
for row in range(2, sheet1.max_row+1):
    quantity = sheet1.cell(row, 2)
    cost = sheet1.cell(row, 3)
    subTotalCell = sheet1.cell(row, 4)
    subTotal = float(quantity.value) * float(cost.value)
    subTotalCell.value = subTotal

pie = PieChart()
print(sheet1.max_row)
labels = Reference(sheet1, min_col=1, min_row=2, max_row=sheet1.max_row)
data = Reference(sheet1, min_col=4, min_row=1, max_row=sheet1.max_row)

pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Expense Report of Dinner"

sheet1.add_chart(pie, 'e2')
wb.save('pythonExcelResults.xlsx')









